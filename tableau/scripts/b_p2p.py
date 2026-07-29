import openpyxl, csv
wb = openpyxl.load_workbook('src/1018f589-P2P_Rollout_Tracker__Monthly_Input_Sheet.xlsx', read_only=True, data_only=True)
ws = wb['P2P Roll-out Tracker']
# Known layout (0-based): 1=Region 2=Market 3=Product 4=Estate 5..7=2023/24/25 8..19=Jan..Dec
MONTHS=['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
VALID_REGION={'AMER','EMEAA','GC','GR CHINA','Global','Greater China'}
VALID_ESTATE={'Managed','Franchised','Managed & Franchised','Managed - N/A'}
rows=[]
for r in ws.iter_rows(min_row=6,max_row=ws.max_row,max_col=28,values_only=True):
    if not r or len(r)<20: continue
    reg,mkt,prod,est = r[1],r[2],r[3],r[4]
    if not isinstance(reg,str) or reg.strip() not in VALID_REGION: continue
    if not isinstance(est,str) or est.strip() not in VALID_ESTATE: continue
    # exclude subtotal rows (Market='All' or Product='All') - facts must be single-grain
    if (isinstance(mkt,str) and mkt.strip()=='All') or (isinstance(prod,str) and prod.strip()=='All'): continue
    reg=reg.strip(); est=est.strip()
    mkt=(mkt or '').strip() if isinstance(mkt,str) else str(mkt or '')
    prod=(prod or '').strip() if isinstance(prod,str) else str(prod or '')
    for off,yr in [(5,2023),(6,2024),(7,2025)]:
        v=r[off]
        if isinstance(v,(int,float)): rows.append([reg,mkt,prod,est,f'{yr}-12-01',yr,'Year End',float(v)])
    for i,mn in enumerate(MONTHS):
        v=r[8+i]
        if isinstance(v,(int,float)): rows.append([reg,mkt,prod,est,f'2026-{i+1:02d}-01',2026,'Monthly',float(v)])
wb.close()
with open('out/Fact_P2P.csv','w',newline='',encoding='utf-8') as f:
    w=csv.writer(f); w.writerow(['region','market','product','estate','month','year','basis','systems']); w.writerows(rows)
print(f"Fact_P2P.csv  {len(rows):,} rows")
