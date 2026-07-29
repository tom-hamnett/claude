import openpyxl, csv
MON={'January':1,'February':2,'March':3,'April':4,'May':5,'June':6,'July':7,'August':8,
     'September':9,'October':10,'November':11,'December':12}
wb=openpyxl.load_workbook('src/8750b514-System_Size_Growth.xlsx',read_only=True,data_only=True)
rows=[]
for sn in wb.sheetnames:
    ws=wb[sn]
    grid=[list(r)+[None]*(14-len(r)) for r in ws.iter_rows(min_row=1,max_row=ws.max_row,max_col=14,values_only=True)]
    hi=None
    for i,r in enumerate(grid):
        if sum(1 for c in r if isinstance(c,str) and c.strip() in MON)>=4: hi=i; break
    if hi is None: continue
    mcols={j:MON[c.strip()] for j,c in enumerate(grid[hi]) if isinstance(c,str) and c.strip() in MON}
    mc0=min(mcols)                      # first month column
    gcol, mcol = mc0-2, mc0-1           # geography, metric sit immediately left
    estate='Managed' if 'Managed' in sn else 'Franchise'
    unit='Rooms' if 'Rooms' in sn else 'Hotels'
    for r in grid[hi+1:]:
        geo = r[gcol].strip() if gcol>=0 and isinstance(r[gcol],str) else None
        met = r[mcol].strip() if isinstance(r[mcol],str) else None
        if not geo or not met: continue
        for j,mn in mcols.items():
            v=r[j]
            if isinstance(v,(int,float)) and not isinstance(v,bool):
                rows.append([estate,unit,geo,met,f'2026-{mn:02d}-01',2026,float(v)])
wb.close()
with open('out/Fact_SystemSize.csv','w',newline='',encoding='utf-8') as f:
    w=csv.writer(f); w.writerow(['estate','unit','geography','metric','month','year','value']); w.writerows(rows)
print(f"Fact_SystemSize.csv  {len(rows):,} rows")
