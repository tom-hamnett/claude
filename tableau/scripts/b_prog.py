import openpyxl, csv
wb=openpyxl.load_workbook('src/14850d0c-Market_Analysis.xlsx',read_only=True,data_only=True)
ws=wb['IHG_Prog_Spend']
rows=[]
it=ws.iter_rows(min_row=1,max_row=ws.max_row,max_col=17,values_only=True)
hdr=list(next(it))
YRS={9:'2021',10:'2022',11:'2023',12:'2024'}
for r in it:
    if not r or not r[0]: continue
    name,code,creg,ctry,brand,rms,l1,l2,od = r[0],r[1],r[2],r[3],r[4],r[5],r[6],r[7],r[8]
    lifecycle = r[13] if len(r)>13 else None
    for j,yr in YRS.items():
        v=r[j] if j<len(r) else None
        if isinstance(v,(int,float)) and v:
            rows.append([code,name,creg,ctry,brand,rms,l1,l2,lifecycle,yr,'Total Spend',float(v)])
    v=r[16] if len(r)>16 else None
    if isinstance(v,(int,float)) and v:
        rows.append([code,name,creg,ctry,brand,rms,l1,l2,lifecycle,'2025','Programme (P2P) Spend',float(v)])
wb.close()
with open('out/Fact_Programme_Spend.csv','w',newline='',encoding='utf-8') as f:
    w=csv.writer(f)
    w.writerow(['hotel_code','hotel_name','company_region','country','brand_code','rooms',
                'category_l1','category_l2','lifecycle_stage','year','measure','spend'])
    w.writerows(rows)
print(f"Fact_Programme_Spend.csv  {len(rows):,} rows")
