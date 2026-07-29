import openpyxl, csv, datetime

def mname(v):
    if isinstance(v,(datetime.datetime,datetime.date)): return v.strftime('%Y-%m-01')
    return None

# ---------- CRF: region x month, all years ----------
wb = openpyxl.load_workbook('src/adde0aeb-CRF_Analysis_20232026.xlsx', read_only=True, data_only=True)
rows=[]
for sn in ['2023','2024','2025','2026']:
    ws=wb[sn]
    grid=[list(r) for r in ws.iter_rows(min_row=1,max_row=8,max_col=16,values_only=True)]
    # header row = the one containing 'Region'
    hdr=None
    for r in grid:
        if r and any(isinstance(c,str) and c.strip()=='Region' for c in r): hdr=r; break
    if not hdr: continue
    ci={i:mname(c) for i,c in enumerate(hdr) if mname(c)}
    for r in grid:
        if not r: continue
        lbl=None
        for c in r:
            if isinstance(c,str) and c.strip() in ('AMER','EMEAA','GC'): lbl=c.strip(); break
        if not lbl: continue
        for i,m in ci.items():
            v=r[i]
            if isinstance(v,(int,float)):
                rows.append([lbl,m,int(sn),round(float(v),2)])
wb.close()
with open('out/Fact_CRF.csv','w',newline='',encoding='utf-8') as f:
    w=csv.writer(f); w.writerow(['region','month','year','crf_usd']); w.writerows(rows)
print(f"Fact_CRF.csv            {len(rows):>7,} rows  (2023-2026, region x month)")

# ---------- P2P: region/market/product/estate x month ----------
wb = openpyxl.load_workbook('src/1018f589-P2P_Rollout_Tracker__Monthly_Input_Sheet.xlsx', read_only=True, data_only=True)
ws = wb['P2P Roll-out Tracker']
grid=[list(r) for r in ws.iter_rows(min_row=1,max_row=ws.max_row,max_col=28,values_only=True)]
hdr=grid[4]  # r5
idx={}
for i,c in enumerate(hdr):
    if isinstance(c,str): idx[c.strip()]=i
months=['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
prows=[]
for r in grid[5:]:
    if not r or not isinstance(r[idx.get('Region',1)],str): continue
    reg=r[idx['Region']]; mkt=r[idx.get('Market',2)]; prod=r[idx.get('Product',3)]; est=r[idx.get('Estate',4)]
    if not reg or reg.strip()=='': continue
    for yr in ['2023','2024','2025']:
        if yr in idx and isinstance(r[idx[yr]],(int,float)):
            prows.append([reg,mkt,prod,est,f'{yr}-12-01',int(yr),'Year End',r[idx[yr]]])
    for mi,mn in enumerate(months,1):
        if mn in idx and isinstance(r[idx[mn]],(int,float)):
            prows.append([reg,mkt,prod,est,f'2026-{mi:02d}-01',2026,'Monthly',r[idx[mn]]])
wb.close()
with open('out/Fact_P2P.csv','w',newline='',encoding='utf-8') as f:
    w=csv.writer(f); w.writerow(['region','market','product','estate','month','year','basis','systems']); w.writerows(prows)
print(f"Fact_P2P.csv            {len(prows):>7,} rows  (region/market/product/estate x month)")
