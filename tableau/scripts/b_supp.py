import openpyxl, csv, datetime
wb=openpyxl.load_workbook('src/d07b0129-Supplier_Metrics_Input_Sheet_RREVSed.xlsx',read_only=True,data_only=True)
ws=wb['Supplier Performance & Risk']
grid=[list(r)+[None]*(25-len(r)) for r in ws.iter_rows(min_row=1,max_row=ws.max_row,max_col=25,values_only=True)]
wb.close()

BLOCKS={2:'Supplier Baseline',28:'Supplier Code of Conduct',34:'Sedex Assessment',60:'EcoVadis',88:'Rapid Ratings'}
NOTES_START=121   # commentary below this
def blk(i):
    cur='Supplier Baseline'
    for r,n in sorted(BLOCKS.items()):
        if i>=r: cur=n
    return cur

# find date columns from any header row
datecols={}
for i,r in enumerate(grid,1):
    ds=[(j,c) for j,c in enumerate(r) if isinstance(c,(datetime.datetime,datetime.date))]
    if len(ds)>=6:
        for j,c in ds: datecols[j]=c.strftime('%Y-%m-01')
# static period columns
STATIC={3:'2025 FY Spend',5:'2025 EoY'}
rows=[]
for i,r in enumerate(grid,1):
    if i>=NOTES_START: break
    rtype = r[0].strip() if isinstance(r[0],str) else None
    metric = r[2].strip() if isinstance(r[2],str) else None
    if not metric: continue
    if metric in BLOCKS.values(): continue          # section header itself
    if metric.startswith('[') or metric.endswith(':'): continue
    for j,per in list(STATIC.items())+list(datecols.items()):
        v=r[j] if j<len(r) else None
        if isinstance(v,(int,float)) and not isinstance(v,bool):
            rows.append([blk(i), rtype or '', metric, per, float(v)])
with open('out/Fact_Supplier.csv','w',newline='',encoding='utf-8') as f:
    w=csv.writer(f); w.writerow(['programme','row_type','metric','period','value']); w.writerows(rows)
print(f"Fact_Supplier.csv  {len(rows):,} rows")
